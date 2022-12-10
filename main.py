from openpyxl import Workbook
from config import new_domain
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from bs4 import BeautifulSoup
from database import create_sql_query, clean_photo
import os
import shutil
import pathlib
import requests
import json
import time
import re


ALPHA = {"а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "j", "з": "z", "и": "i",
         "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t",
         "у": "u", "ф": "f", "х": "x", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sh", "ъ": "", "ы": "i", "ь": "",
         "э": "e", "ю": "yu", "я": "ya"}
index_photo = 1

workbook = Workbook()
oc_product = workbook.active
oc_product.title = "oc_product"
oc_product_description = workbook.create_sheet(title="oc_product_description")
oc_product_image = workbook.create_sheet(title="oc_product_image")
oc_product_to_category = workbook.create_sheet(title="oc_product_to_category")
oc_seo_url = workbook.create_sheet(title='oc_seo_url')

ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
options = webdriver.ChromeOptions()
options.add_argument(f"user-agent={ua}")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--headless")
timeout = 300


def new_excel():
    global oc_product
    global oc_product_description
    global oc_product_image
    global oc_product_to_category
    global oc_seo_url

    # oc_product page
    oc_product["A1"] = "product_id"
    oc_product["B1"] = "quantity"
    oc_product["C1"] = "stock_status_id"
    oc_product["D1"] = "image"
    oc_product["E1"] = "shipping"
    oc_product["F1"] = "price"
    oc_product["G1"] = "model"

    # oc_product_description page
    oc_product_description["A1"] = "product_id"
    oc_product_description["B1"] = "language_id"
    oc_product_description["C1"] = "name"
    oc_product_description["D1"] = "description"
    oc_product_description["E1"] = "meta_title"
    oc_product_description["F1"] = "meta_description"

    # oc_product_image page
    oc_product_image["A1"] = "product_image_id"
    oc_product_image["B1"] = "product_id"
    oc_product_image["C1"] = "image"
    oc_product_image["D1"] = "sort_order"
    oc_product_image["E1"] = "image_path"

    # oc_product_to_category page
    oc_product_to_category["A1"] = "product_id"
    oc_product_to_category["B1"] = "category_id"
    oc_product_to_category["C1"] = "product_name"

    # oc_seo_url page
    oc_seo_url["A1"] = "seo_url_id"
    oc_seo_url["B1"] = "store_id"
    oc_seo_url["C1"] = 'language_id'
    oc_seo_url["D1"] = "query"
    oc_seo_url["E1"] = "keyword"


def write_data(data, index, mode):
    global oc_product
    global oc_product_description
    global oc_product_image
    global oc_product_to_category
    global oc_seo_url
    global index_photo

    oc_product[f"C{index}"] = 8
    oc_product[f"D{index}"] = data["photos"][0]["correct_url"]
    oc_product[f"E{index}"] = 1
    oc_product[f"F{index}"] = data["full_price"]

    oc_product_description[f"A{index}"] = f"=oc_product!A{index}"
    oc_product_description[f"B{index}"] = 2
    oc_product_description[f"C{index}"] = data["title"]
    oc_product_description[f"D{index}"] = data["description"]

    for photo in data["photos"]:
        index_photo += 1
        oc_product_image[f"B{index_photo}"] = f"=oc_product!A{index}"
        oc_product_image[f"C{index_photo}"] = photo["correct_url"]
        oc_product_image[f"D{index_photo}"] = 0
        if mode == "photo":
            oc_product_image[f"E{index_photo}"] = str(pathlib.Path(os.getcwd(), photo["path"]))
        else:
            oc_product_image[f"E{index_photo}"] = 'Данные были собраны в режиме "Без Фото"'

    oc_product_to_category[f"A{index}"] = f"=oc_product!A{index}"
    oc_product_to_category[f"C{index}"] = data["title"]

    oc_seo_url[f"B{index}"] = 0
    oc_seo_url[f"C{index}"] = 2
    oc_seo_url[f"D{index}"] = f'="product_id="&oc_product!A{index}'
    oc_seo_url[f"E{index}"] = create_seo_url(data["title"])


def create_seo_url(title):
    result = list()
    sub_result = list()
    for symbol in title.lower():
        if symbol.isalnum():
            sub_result.append(symbol)
        else:
            sub_result.append("-")
    for symbol in sub_result:
        if symbol == "-" or symbol.isdigit():
            result.append(symbol)
        else:
            if symbol in ALPHA.keys():
                result.append(ALPHA[symbol])
            else:
                result.append(symbol)
    result = "".join(result).replace("--", "-")
    if result[0] == "-":
        result = result[1:]
    if result[-1] == "-":
        result = result[:-1]
    return result


def create_photo_dir():
    if os.path.isdir("photos"):
        shutil.rmtree("photos")
        os.mkdir("photos")
    else:
        os.mkdir("photos")
    print("[INFO] Папка для фотографий photos успешно создана")


def get_photo_data(index, title, mode):
    correct_url = f'{new_domain}/{create_seo_url(title=title)}{index}.jpg'
    result = {"correct_url": correct_url}
    if mode == "photo":
        path = pathlib.Path("photos", f"{create_seo_url(title=title)}{index}.jpg")
        result["path"] = path
    return result


def download_photos(photos, title):
    index = 0
    result = list()
    for photo in photos:
        index += 1
        response = requests.get(photo)
        path = pathlib.Path("photos", f"{create_seo_url(title)}{index}.jpg")
        with open(path, "wb") as file:
            file.write(response.content)
        result.append(get_photo_data(title=title, index=index, mode="photo"))
    return result


def get_product_ids():
    browser = webdriver.Chrome(options=options)
    browser.set_page_load_timeout(time_to_wait=timeout)
    wait = WebDriverWait(browser, timeout)
    try:
        url = "https://flowwow.com/shop/flower-bag/"
        browser.get(url=url)
        time.sleep(2)
        wait.until(ec.visibility_of_element_located((By.CLASS_NAME, "store_readmore")))
        read_more_buttons = browser.find_elements(By.LINK_TEXT, "Показать ещё")
        for read_more_button in read_more_buttons:
            wait.until(ec.element_to_be_clickable((By.CLASS_NAME, "store_readmore")))
            browser.execute_script("arguments[0].scrollIntoView();", read_more_button)
            browser.execute_script("arguments[0].click();", read_more_button)
            time.sleep(1)
        last_block = browser.find_elements(By.TAG_NAME, "div")[-1]
        browser.execute_script("arguments[0].scrollIntoView();", last_block)
        time.sleep(2)
        response = browser.page_source
        bs_object = BeautifulSoup(response, "lxml")
        products = bs_object.find_all(name="a", class_=re.compile(r"shop-product js-product-popu\w*"))
        product_ids = set([product["data-id"] for product in products])
        print(f"[INFO] В магазине найдено {len(product_ids)} уникальных товаров")
        return product_ids
    finally:
        browser.close()
        browser.quit()


def get_data(product_ids, mode):
    browser = webdriver.Chrome(options=options)
    browser.set_page_load_timeout(time_to_wait=timeout)
    index_field = 1
    try:
        for product_id in product_ids:
            print(f"[INFO] Собираем данные о товаре номер {product_id}")
            index_field += 1
            url = f"https://flowwow.com/moscow/data/getProductInfo/?id={product_id}&from=direct&lang=ru&currency=RUB"
            browser.get(url=url)
            response = browser.page_source
            bs_object = BeautifulSoup(response, "lxml")
            json_object = json.loads(bs_object.text)["data"]

            if "base" in json_object.keys():
                full_price = json_object["base"]
            else:
                full_price = json_object["cost"]

            full_info = json_object["fullInfo"]
            full_info = full_info.replace("\n", "").replace("\\", "")
            bs_object = BeautifulSoup(full_info, "lxml")
            title = bs_object.find(name="div", class_='pp-title').text.strip().replace('"', "")

            composition = bs_object.find(name="div", class_="product-desc-line").find(name="div", class_='desc')
            if composition is not None:
                composition = composition.text.strip().split(".")
                composition = ", ".join(element.strip() for element in composition)
                composition = composition.replace("Показать ещё", "")
                composition = f"Состав: {composition}"
                if len(composition) < 1:
                    composition = ""
            else:
                composition = ""

            size = bs_object.find_all(name="div", class_="product-desc-line")
            if len(size) > 1:
                size = size[1].find(name="div", class_='desc')
                if size is not None:
                    size = size.text.strip().replace(" ", "")
                    size_index = size.find("Ш")
                    size = size[:size_index] + "; " + size[size_index:]
                    size = f"Размер: {size}"
                    if len(size) < 1:
                        size = ""
                else:
                    size = ""
            else:
                size = ""

            description = bs_object.find(name="div", class_='product-describe')
            if description is not None:
                description = description.text.strip().replace('"', "")
                if len(description) < 1:
                    description = ""
            else:
                description = ""

            description = "\n".join([description, composition, size])
            seo_url = create_seo_url(title)

            photo_objects = json_object["photos"]
            photos = list()
            for photo_object in photo_objects:
                if "img" in photo_object.keys():
                    photos.append(photo_object["img"].replace('"', ""))

            if mode == "photo":
                photos = download_photos(photos=photos, title=title)
            else:
                result_photos = list()
                index = 0
                for _ in photos:
                    index += 1
                    result_photo = get_photo_data(title=title, index=index, mode="only_text")
                    result_photos.append(result_photo)
                photos = result_photos
            result = {"title": title, "full_price": full_price,
                      "description": description, "seo_url": seo_url, "photos": photos}

            write_data(data=result, index=index_field, mode=mode)
            print(f"[INFO] Собрано и записано {index_field - 1}/{len(product_ids)} товаров")
    finally:
        browser.close()
        browser.quit()


def parsing(mode):
    if mode == "photo":
        program_time = "20 минут"
    else:
        program_time = "5 минут"
    print("[INFO] Программа запущена")
    print(f"[INFO] Идет сбор информации из магазина Flower Bag. Это займет не более {program_time}")
    start_time = time.time()
    product_ids = get_product_ids()
    get_data(product_ids=product_ids, mode=mode)
    print("[INFO] Сбор информации закончен")
    stop_time = time.time()
    total_time = stop_time - start_time
    print("[INFO] Программа завершена")
    print("[INFO] Результат работы парсера хранится в файле result.csv в той же папке, где хранится программа")
    print(f"[INFO] На работу программы ушло {total_time} секунд")


def main():
    global workbook
    print("[INFO] Привет! Я автоматизированный помощник по сбору и записи данных")
    print("[INFO] У меня есть подробная инструкция в файле ИНСТРУКЦИЯ.txt и я очень советую сначала ее прочитать")
    print('[INFO] Чтобы включить режим парсинга, введи 1 в консоль. Для выбора парсинга без фото введите 2.', end=" ")
    print("Для выбора режима записи введите 3")
    while True:
        mode = input("[INPUT] Выбери режим работы: >>> ")
        if mode == "1":
            new_excel()
            create_photo_dir()
            parsing(mode="photo")
            workbook.save("result.xlsx")
            break
        elif mode == "2":
            new_excel()
            parsing(mode="only_text")
            workbook.save("result.xlsx")
            break
        elif mode == "3":
            clean_photo()
            create_sql_query()
            break
        else:
            print("[ERROR] Прости, похоже ты ввел не ту команду. Попробуй еще раз")


if __name__ == "__main__":
    main()
