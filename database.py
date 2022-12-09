from config import db_name
import openpyxl
import os
import pathlib


def clean_photo():
    print("[INFO] Идет зачистка ненужных фотографий из системы")
    workbook = openpyxl.load_workbook("result.xlsx")
    oc_product_image = workbook["oc_product_image"]

    dir_photos = os.listdir(path=pathlib.Path(os.getcwd(), "photos"))
    excel_photos = list()
    for index in range(2, oc_product_image.max_row + 1):
        excel_photos.append(oc_product_image[f"E{index}"].value)

    undelete_list = list()
    for dir_photo in dir_photos:
        for excel_photo in excel_photos:
            if excel_photo is not None:
                if dir_photo == excel_photo.split("\\")[-1]:
                    undelete_list.append(dir_photo)

    for dir_photo in dir_photos:
        if dir_photo not in undelete_list:
            os.remove(path=pathlib.Path(os.getcwd(), "photos", str(dir_photo)))
            print(f"[INFO] Фотография {str(dir_photo)} была удалена из памяти компьютера")
    print("[INFO] Зачистка закончена")


def get_clean_value(workbook, formula):
    item_list = formula[1:].split("!")
    page = item_list[0]
    cell = item_list[1]
    result = workbook[page][cell].value
    return result


def get_query_value(workbook, formula):
    item_list = formula.split("&")
    sub_formula = item_list[1].split("!")
    page = sub_formula[0]
    cell = sub_formula[1]
    result = f"product_id={workbook[page][cell].value}"
    return result


def create_sql_query():
    print("[INFO] Формируем SQL-запрос")
    workbook = openpyxl.load_workbook("result.xlsx")
    oc_product = workbook["oc_product"]
    oc_product_description = workbook["oc_product_description"]
    oc_product_image = workbook["oc_product_image"]
    oc_product_to_category = workbook["oc_product_to_category"]
    oc_seo_url = workbook["oc_seo_url"]

    # select oc_product
    oc_product_query = " ".join(["INSERT INTO oc_product (`product_id`, `quantity`,",
                                 "`stock_status_id`, `image`, `shipping`, `price`) VALUES\n"])
    values_queries = list()
    for index in range(2, oc_product.max_row + 1):
        product_id = oc_product[f"A{index}"].value
        quantity = oc_product[f"B{index}"].value
        stock_status_id = oc_product[f"C{index}"].value
        image = oc_product[f"D{index}"].value
        shipping = oc_product[f"E{index}"].value
        price = oc_product[f"F{index}"].value
        values_query = f"""({product_id}, {quantity}, {stock_status_id}, "{image}", {shipping}, {price})"""
        values_queries.append(values_query)
    values_queries = ",\n".join(values_queries)
    oc_product_query = f"{oc_product_query}{values_queries};"

    # select oc_product_description
    oc_product_description_query = " ".join(["INSERT INTO oc_product_description",
                                             "(`product_id`, `language_id`, `name`, `description`,",
                                             "`meta_title`, `meta_description`) VALUES\n"])
    values_queries = list()
    for index in range(2, oc_product_description.max_row + 1):
        product_id = get_clean_value(workbook=workbook, formula=oc_product_description[f"A{index}"].value)
        language_id = oc_product_description[f"B{index}"].value
        name = str(oc_product_description[f"C{index}"].value).replace('"', "'")
        description = str(oc_product_description[f"D{index}"].value).replace('"', "'")
        meta_title = oc_product_description[f"E{index}"].value
        meta_description = oc_product_description[f"F{index}"].value
        values_query = " ".join([f'({product_id}, {language_id}, "{name}",',
                                 f'"{description}", "{meta_title}", "{meta_description}")'])
        values_queries.append(values_query)
    values_queries = ",\n".join(values_queries)
    oc_product_description_query = f"{oc_product_description_query}{values_queries};"

    # select oc_product_image
    oc_product_image_query = " ".join(["INSERT INTO oc_product_image (`product_image_id`,",
                                      "`product_id`, `image`, `sort_order`) VALUES\n"])
    values_queries = list()
    for index in range(2, oc_product_image.max_row + 1):
        if oc_product_image[f"B{index}"].value is not None:
            product_image_id = oc_product_image[f"A{index}"].value
            product_id = get_clean_value(workbook=workbook, formula=oc_product_image[f"B{index}"].value)
            image = oc_product_image[f"C{index}"].value
            sort_order = oc_product_image[f"D{index}"].value
            values_query = f"({product_image_id}, {product_id}, '{image}', {sort_order})"
            values_queries.append(values_query)
    values_queries = ",\n".join(values_queries)
    oc_product_image_query = f"{oc_product_image_query}{values_queries};"

    # select oc_product_to_category
    oc_product_to_category_query = "INSERT INTO oc_product_to_category (`product_id`, `category_id`) VALUES\n"
    values_queries = list()
    for index in range(2, oc_product_to_category.max_row + 1):
        product_id = get_clean_value(workbook=workbook, formula=oc_product_to_category[f"A{index}"].value)
        category_id = oc_product_to_category[f"B{index}"].value
        values_query = f"({product_id}, {category_id})"
        values_queries.append(values_query)
    values_queries = ",\n".join(values_queries)
    oc_product_to_category_query = f"{oc_product_to_category_query}{values_queries};"

    # select oc_seo_url
    oc_seo_url_query = "INSERT INTO oc_seo_url (`seo_url_id`, `store_id`, `language_id`, `query`, `keyword`) VALUES\n"
    values_queries = list()
    for index in range(2, oc_seo_url.max_row + 1):
        seo_url_id = oc_seo_url[f"A{index}"].value
        store_id = oc_seo_url[f"B{index}"].value
        language_id = oc_seo_url[f"C{index}"].value
        query = get_query_value(workbook=workbook, formula=oc_seo_url[f"D{index}"].value)
        keyword = oc_seo_url[f"E{index}"].value
        values_query = f"({seo_url_id}, {store_id}, {language_id}, '{query}', '{keyword}')"
        values_queries.append(values_query)
    values_queries = ",\n".join(values_queries)
    oc_seo_url_query = f"{oc_seo_url_query}{values_queries};"

    result = "\n".join([f"USE {db_name};", oc_product_query, oc_product_description_query,
                        oc_product_image_query, oc_product_to_category_query, oc_seo_url_query])
    with open("sql.txt", "w", encoding="utf-8") as file:
        file.write(result)
    print("[INFO] SQL-запрос успешно сгенерирован и сохранен в файле sql.txt")


if __name__ == "__main__":
    clean_photo()
